import csv
import json
import os
import uuid
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, List, Tuple
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from config import get_config


def _export_dir() -> str:
    cfg = get_config()
    base = os.path.abspath(cfg.EXPORT_BASE_DIR)
    os.makedirs(base, exist_ok=True)
    return base


def _table(data):
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    return table


def _write_pdf(payload: Dict[str, Any], title: str, company_name: str) -> str:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Paragraph(company_name, styles["Heading2"])]
    elements.append(Paragraph(f"Generated: {datetime.utcnow().isoformat()} UTC", styles["Normal"]))
    elements.append(Spacer(1, 12))

    kpi = payload.get("kpi_summary", {})
    elements.append(Paragraph("KPI Summary", styles["Heading3"]))
    elements.append(_table([["Metric", "Value"], ["OEE", str(kpi.get("oee", 0))], ["Availability", str(kpi.get("availability", 0))], ["Performance", str(kpi.get("performance", 0))], ["Quality", str(kpi.get("quality", 0))]]))
    elements.append(Spacer(1, 10))

    health = payload.get("health_overview", {})
    elements.append(Paragraph("Health Overview", styles["Heading3"]))
    elements.append(_table([["Avg Health", str(health.get("avg_health", 0))], ["Critical Machines", str(health.get("critical_machines", 0))]]))
    elements.append(Spacer(1, 10))

    fin = payload.get("financial_projection", {})
    elements.append(Paragraph("Financial Projection", styles["Heading3"]))
    elements.append(_table([["Projected Downtime Cost", str(fin.get("projected_downtime_cost", 0))], ["Projected Revenue Loss", str(fin.get("projected_revenue_loss", 0))], ["Total Risk Exposure", str(fin.get("total_risk_exposure", 0))]]))
    elements.append(Spacer(1, 10))

    esg = payload.get("esg", {})
    elements.append(Paragraph("ESG", styles["Heading3"]))
    elements.append(_table([["Total Energy (kWh)", str(esg.get("total_energy_kwh", 0))], ["Sustainability Score", str(esg.get("sustainability_score", 0))], ["Carbon Proxy (kg)", str(esg.get("carbon_proxy_kg", 0))]]))
    elements.append(Spacer(1, 10))

    ai_summary = payload.get("ai_summary", {})
    elements.append(Paragraph("AI Executive Summary", styles["Heading3"]))
    elements.append(Paragraph(ai_summary.get("executive_summary", ""), styles["Normal"]))
    for heading, key in (("Key Risks", "key_risks"), ("Performance Gaps", "performance_gaps"), ("Strategic Recommendations", "strategic_recommendations")):
        items = ai_summary.get(key) or []
        if items:
            elements.append(Paragraph(heading, styles["Heading4"]))
            for item in items:
                elements.append(Paragraph(f"- {item}", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    filename = f"report-{uuid.uuid4().hex}.pdf"
    path = os.path.join(_export_dir(), filename)
    with open(path, "wb") as f:
        f.write(buffer.read())
    return path


def _write_excel(payload: Dict[str, Any], title: str) -> str:
    wb = Workbook()
    ws_kpi = wb.active
    ws_kpi.title = "KPI"
    ws_kpi.append([title])
    ws_kpi.append(["Metric", "Value"])
    kpi = payload.get("kpi_summary", {})
    for key in ("oee", "availability", "performance", "quality"):
        ws_kpi.append([key, kpi.get(key, 0)])

    ws_fin = wb.create_sheet("Financial")
    fin = payload.get("financial_projection", {})
    ws_fin.append(["Metric", "Value"])
    for key in ("projected_downtime_cost", "projected_revenue_loss", "total_risk_exposure"):
        ws_fin.append([key, fin.get(key, 0)])

    ws_pred = wb.create_sheet("AI Predictions")
    ws_pred.append(["Machine", "Plant", "Failure %", "RUL Hours", "Risk", "Created"])
    for row in payload.get("prediction_outlook", []):
        ws_pred.append([
            row.get("machine_id"),
            row.get("plant_id"),
            row.get("failure_probability"),
            row.get("remaining_useful_life_hours"),
            row.get("risk_level"),
            row.get("created_at"),
        ])

    ws_esg = wb.create_sheet("ESG")
    esg = payload.get("esg", {})
    ws_esg.append(["Total Energy (kWh)", esg.get("total_energy_kwh", 0)])
    ws_esg.append(["Sustainability Score", esg.get("sustainability_score", 0)])
    ws_esg.append(["Carbon Proxy (kg)", esg.get("carbon_proxy_kg", 0)])

    ws_ai = wb.create_sheet("AI Summary")
    ai = payload.get("ai_summary", {})
    ws_ai.append(["Executive Summary", ai.get("executive_summary", "")])
    ws_ai.append(["Key Risks"])
    for item in ai.get("key_risks", []) or []:
        ws_ai.append([item])
    ws_ai.append(["Performance Gaps"])
    for item in ai.get("performance_gaps", []) or []:
        ws_ai.append([item])
    ws_ai.append(["Strategic Recommendations"])
    for item in ai.get("strategic_recommendations", []) or []:
        ws_ai.append([item])

    filename = f"report-{uuid.uuid4().hex}.xlsx"
    path = os.path.join(_export_dir(), filename)
    wb.save(path)
    return path


def _write_json(payload: Dict[str, Any]) -> str:
    filename = f"report-{uuid.uuid4().hex}.json"
    path = os.path.join(_export_dir(), filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return path


def _write_csv(payload: Dict[str, Any]) -> str:
    filename = f"report-{uuid.uuid4().hex}.csv"
    path = os.path.join(_export_dir(), filename)
    rows: List[Tuple[str, str]] = []
    kpi = payload.get("kpi_summary", {})
    for key, value in kpi.items():
        rows.append((f"kpi_summary.{key}", str(value)))
    health = payload.get("health_overview", {})
    for key, value in health.items():
        rows.append((f"health_overview.{key}", str(value)))
    fin = payload.get("financial_projection", {})
    for key, value in fin.items():
        rows.append((f"financial_projection.{key}", str(value)))
    esg = payload.get("esg", {})
    for key, value in esg.items():
        rows.append((f"esg.{key}", str(value)))
    ai = payload.get("ai_summary", {})
    rows.append(("ai_summary.executive_summary", ai.get("executive_summary", "")))
    for key in ("key_risks", "performance_gaps", "strategic_recommendations"):
        items = ai.get(key) or []
        rows.append((f"ai_summary.{key}", " | ".join([str(i) for i in items])))

    with open(path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["section", "value"])
        writer.writerows(rows)
    return path


def export_report(payload: Dict[str, Any], export_format: str, title: str, company_name: str) -> str:
    fmt = export_format.upper()
    if fmt == "PDF":
        return _write_pdf(payload, title, company_name)
    if fmt in {"XLS", "XLSX", "EXCEL"}:
        return _write_excel(payload, title)
    if fmt == "CSV":
        return _write_csv(payload)
    return _write_json(payload)
